from __future__ import annotations

import re
from html.parser import HTMLParser
from datetime import date, datetime
from pathlib import Path
from typing import Dict, List, Optional, Set, Tuple

from .calendar_utils import is_duty_day, is_holiday_or_weekend, korean_holidays, month_dates, weekday_ko
from .models import OFF, SHIFT_DAY, SHIFT_DUTY, SHIFT_GY, SHIFT_GY_REST, SHIFT_SWING, Employee, ScheduleMap, ScheduleResult
from .schedule_utils import expand_gy_blocks
from .stats import STAT_HEADERS, averages, compute_stats
from .validation import validate_schedule
from .models import ShiftRules

SHIFT_FILLS = {
    SHIFT_DAY: "FFF2CC",
    SHIFT_SWING: "D9EAD3",
    SHIFT_GY: "D9E2F3",
    SHIFT_DUTY: "FCE4D6",
    SHIFT_GY_REST: "E7E6E6",
    OFF: "FFFFFF",
    "": "FFFFFF",
}


def parse_date(value, default_year: Optional[int] = None, default_month: Optional[int] = None) -> Optional[date]:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%Y.%m.%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    if default_year and default_month:
        try:
            day = int(text)
            return date(default_year, default_month, day)
        except ValueError:
            pass
    return None


def parse_unavailable(text: object, default_year: Optional[int] = None, default_month: Optional[int] = None) -> Set[date]:
    if not text:
        return set()
    result: Set[date] = set()
    for part in str(text).replace(";", ",").replace(" ", ",").split(","):
        d = parse_date(part.strip(), default_year, default_month)
        if d:
            result.add(d)
    return result


def normalize_shift_code(value: object, work_date: Optional[date] = None, holidays: Optional[Set[date]] = None) -> str:
    text = "" if value is None else str(value).strip()
    text = re.sub(r"\s+", "", text)
    text = re.sub(r"[\u200b\u200c\u200d\ufeff]", "", text)
    text = text.replace("／", "/").replace("\\", "/").replace("⁄", "/")
    if not text:
        return OFF
    upper = text.upper()
    if upper in ("D", "DAY", "데이"):
        return SHIFT_DAY
    if upper in ("S", "SW", "SWING", "스윙"):
        return SHIFT_SWING
    if ("G" in upper and ("지근" in text or "지금" in text)) or text in ("지근", "지금", "야간"):
        return SHIFT_GY
    if upper in ("G", "GY"):
        if work_date and is_duty_day(work_date):
            return SHIFT_DUTY
        return SHIFT_GY
    if text in ("당직", "주말당직"):
        return SHIFT_DUTY
    if text in ("지휴", "GY휴", "G휴", "야휴"):
        return SHIFT_GY_REST
    if text in ("휴", "휴무", "OFF", "오프", "-", "."):
        return OFF
    return text


def _looks_like_shift_code(value: object) -> bool:
    text = "" if value is None else str(value).strip()
    if not text:
        return True
    normalized = normalize_shift_code(text)
    return normalized in {OFF, SHIFT_DAY, SHIFT_SWING, SHIFT_GY, SHIFT_DUTY, SHIFT_GY_REST}


def _header_day(value: object) -> Optional[int]:
    if value is None:
        return None
    if isinstance(value, (date, datetime)):
        return value.day
    text = str(value).strip()
    if not text:
        return None
    match = re.search(r"\d{1,2}", text)
    if not match:
        return None
    day = int(match.group())
    if 1 <= day <= 31:
        return day
    return None



class _HTMLTableParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.rows: List[List[dict]] = []
        self.class_styles: Dict[str, str] = {}
        self._current_row: Optional[List[dict]] = None
        self._current_cell: Optional[dict] = None
        self._capture = False
        self._capture_style = False
        self._style_text = ""

    def handle_starttag(self, tag: str, attrs) -> None:
        attrs_dict = {k.lower(): (v or "") for k, v in attrs}
        tag = tag.lower()
        if tag == "style":
            self._capture_style = True
            self._style_text = ""
        elif tag == "tr":
            self._current_row = []
        elif tag in ("td", "th") and self._current_row is not None:
            self._current_cell = {
                "text": "",
                "style": attrs_dict.get("style", ""),
                "bgcolor": attrs_dict.get("bgcolor", ""),
                "class": attrs_dict.get("class", ""),
            }
            self._capture = True

    def handle_data(self, data: str) -> None:
        if self._capture_style:
            self._style_text += data
        elif self._capture and self._current_cell is not None:
            self._current_cell["text"] += data

    def handle_endtag(self, tag: str) -> None:
        tag = tag.lower()
        if tag == "style":
            self._capture_style = False
            self.class_styles.update(_parse_css_class_styles(self._style_text))
            self._style_text = ""
        elif tag in ("td", "th") and self._current_row is not None and self._current_cell is not None:
            self._current_cell["text"] = " ".join(self._current_cell["text"].split())
            self._current_row.append(self._current_cell)
            self._current_cell = None
            self._capture = False
        elif tag == "tr" and self._current_row is not None:
            if self._current_row:
                self.rows.append(self._current_row)
            self._current_row = None


def _parse_css_class_styles(css: str) -> Dict[str, str]:
    result: Dict[str, str] = {}
    for match in re.finditer(r"\.([A-Za-z0-9_-]+)\s*\{([^}]*)\}", css or "", flags=re.S):
        result[match.group(1)] = " ".join(match.group(2).split())
    return result


def _hex_to_rgb(value: str) -> Optional[tuple[int, int, int]]:
    value = value.strip().strip('"\'')
    if not value:
        return None
    embedded = re.search(r"#?([0-9a-fA-F]{6})", value)
    if embedded:
        value = embedded.group(1)
    if value.startswith("#"):
        value = value[1:]
    if len(value) == 3:
        value = "".join(ch * 2 for ch in value)
    if len(value) != 6:
        return None
    try:
        return int(value[0:2], 16), int(value[2:4], 16), int(value[4:6], 16)
    except ValueError:
        return None


def _css_rgb_to_tuple(value: str) -> Optional[tuple[int, int, int]]:
    match = re.search(r"rgba?\((\d+)\s*,\s*(\d+)\s*,\s*(\d+)", value, flags=re.I)
    if not match:
        return None
    return tuple(int(match.group(i)) for i in range(1, 4))  # type: ignore[return-value]


def _cell_color_from_html_cell(cell: dict) -> Optional[tuple[int, int, int]]:
    style = cell.get("style", "") or ""
    class_style = cell.get("class_style", "") or ""
    if class_style:
        style = f"{class_style};{style}"
    bgcolor = cell.get("bgcolor", "") or ""
    if bgcolor:
        rgb = _hex_to_rgb(bgcolor)
        if rgb:
            return rgb
    # Excel/Sheets usually use background, background-color, or mso-pattern.
    for pattern in (
        r"background(?:-color)?\s*:\s*([^;]+)",
        r"mso-pattern\s*:[^#;]*(#[0-9a-fA-F]{6})",
    ):
        match = re.search(pattern, style, flags=re.I)
        if not match:
            continue
        raw = match.group(1).strip()
        rgb = _css_rgb_to_tuple(raw) or _hex_to_rgb(raw)
        if rgb:
            return rgb
    return None


def is_gray_rgb(rgb: Optional[tuple[int, int, int]]) -> bool:
    if rgb is None:
        return False
    r, g, b = rgb
    if (r, g, b) in ((255, 255, 255), (0, 0, 0)):
        return False
    return max(r, g, b) - min(r, g, b) <= 22 and 65 <= (r + g + b) / 3 <= 240


def is_blue_rgb(rgb: Optional[tuple[int, int, int]]) -> bool:
    if rgb is None:
        return False
    r, g, b = rgb
    # Excel's dark/medium blue fills vary by theme. Treat clearly blue cells as
    # G/지근 continuation marks when they have no text.
    if b < 70:
        return False
    return b >= r + 25 and b >= g + 10


def parse_html_table(html: str) -> List[List[dict]]:
    parser = _HTMLTableParser()
    parser.feed(html or "")
    if parser.class_styles:
        for row in parser.rows:
            for cell in row:
                class_names = str(cell.get("class", "")).split()
                styles = [parser.class_styles[name] for name in class_names if name in parser.class_styles]
                if styles:
                    cell["class_style"] = ";".join(styles)
    return parser.rows


def _rows_to_text_matrix(rows: List[List[dict]]) -> List[List[str]]:
    return [[str(cell.get("text", "")).strip() for cell in row] for row in rows]


def _rows_to_schedule_matrix(rows: List[List[dict]]) -> List[List[str]]:
    matrix: List[List[str]] = []
    for row in rows:
        values: List[str] = []
        for cell in row:
            text = str(cell.get("text", "")).strip()
            if not text and is_blue_rgb(_cell_color_from_html_cell(cell)):
                text = SHIFT_GY
            values.append(text)
        matrix.append(values)
    return matrix


def parse_schedule_from_clipboard(text: str, html: str, year: int, month: int, rules: Optional[ShiftRules] = None) -> ScheduleResult:
    html_rows = parse_html_table(html)
    if html_rows:
        tsv = "\n".join("\t".join(row) for row in _rows_to_schedule_matrix(html_rows))
        return parse_schedule_from_tsv(tsv, year, month, rules)
    return parse_schedule_from_tsv(text, year, month, rules)


def parse_unavailable_from_clipboard(text: str, html: str, year: int, month: int) -> Dict[str, Set[date]]:
    html_rows = parse_html_table(html)
    if not html_rows:
        raise ValueError("클립보드에 셀 색상 정보가 없습니다. 엑셀에서 범위를 복사한 뒤 바로 붙여넣으세요.")

    rows = html_rows
    header_index = 0
    name_col = 0
    id_col = 1
    for idx, row in enumerate(rows[:8]):
        values = [str(c.get("text", "")).strip().replace(" ", "").lower() for c in row]
        if any(v in ("성명", "이름", "name") for v in values):
            header_index = idx
            for col, v in enumerate(values):
                if v in ("성명", "이름", "name"):
                    name_col = col
                if v in ("사번", "직번", "id", "employeeid", "employee_id"):
                    id_col = col
            break

    valid_dates = {d.day: d for d in month_dates(year, month)}
    day_cols: Dict[int, int] = {}
    for col, cell in enumerate(rows[header_index]):
        day = _header_day(cell.get("text", ""))
        if day in valid_dates:
            day_cols[day] = col
    if not day_cols:
        for offset, d in enumerate(month_dates(year, month), start=2):
            day_cols[d.day] = offset

    result: Dict[str, Set[date]] = {}
    for row in rows[header_index + 1:]:
        if name_col >= len(row):
            continue
        name = str(row[name_col].get("text", "")).strip()
        if not name:
            continue
        employee_no = str(row[id_col].get("text", "")).strip() if id_col < len(row) else ""
        dates: Set[date] = set()
        for day, col in day_cols.items():
            if col < len(row) and is_gray_rgb(_cell_color_from_html_cell(row[col])):
                dates.add(valid_dates[day])
        if dates:
            key = f"{name}|{employee_no}"
            result[key] = dates
            if employee_no:
                result[employee_no] = dates
            result[name] = dates
    return result

def parse_schedule_from_tsv(text: str, year: int, month: int, rules: Optional[ShiftRules] = None) -> ScheduleResult:
    rows = [[cell.strip() for cell in line.split("\t")] for line in text.splitlines() if line.strip()]
    if not rows:
        raise ValueError("붙여넣은 표가 비어 있습니다.")

    header_index: Optional[int] = None
    name_col = 0
    id_col = 1
    day_cols: Dict[int, int] = {}

    for idx, row in enumerate(rows[:8]):
        lowered = [c.replace(" ", "").lower() for c in row]
        if any(c in ("성명", "이름", "name") for c in lowered):
            header_index = idx
            for col, cell in enumerate(lowered):
                if cell in ("성명", "이름", "name"):
                    name_col = col
                if cell in ("사번", "직번", "id", "employeeid", "employee_id"):
                    id_col = col
            break

    valid_dates = {d.day: d for d in month_dates(year, month)}

    if header_index is not None:
        header = rows[header_index]
        for col, cell in enumerate(header):
            day = _header_day(cell)
            if day in valid_dates and col not in (name_col, id_col):
                day_cols[day] = col

    if not day_cols:
        # Header might be omitted because the user clicked below "성명" and
        # copied only data rows. In that case do NOT inspect the first employee
        # row for day numbers: numeric employee IDs such as "21..." would be
        # misread as the 21st day and 사번 would appear under 21일.
        first_data_row = next((row for row in rows if row and row[0].strip()), [])
        has_employee_id_column = len(first_data_row) >= 2 and not _looks_like_shift_code(first_data_row[1])
        id_col = 1 if has_employee_id_column else -1
        first_day_col = 2 if has_employee_id_column else 1
        for offset, d in enumerate(month_dates(year, month), start=first_day_col):
            day_cols[d.day] = offset

    holidays = korean_holidays(year)
    employees: List[Employee] = []
    schedule: ScheduleMap = {d: {} for d in month_dates(year, month)}
    seen_keys: Dict[str, int] = {}

    data_start = (header_index + 1) if header_index is not None else 0
    for row in rows[data_start:]:
        if name_col >= len(row) or not row[name_col].strip():
            continue
        name = row[name_col].strip()
        employee_id = row[id_col].strip() if id_col >= 0 and id_col < len(row) else ""
        base_key = f"{name}|{employee_id}"
        seen_keys[base_key] = seen_keys.get(base_key, 0) + 1
        if seen_keys[base_key] > 1:
            # Same name+employee number duplicated in source. Keep both rows visible
            # instead of overwriting the first row in the schedule map.
            employee_id = f"{employee_id}#{seen_keys[base_key]}" if employee_id else f"row{seen_keys[base_key]}"
        emp = Employee(name=name, employee_id=employee_id)
        employees.append(emp)
        for d in month_dates(year, month):
            col = day_cols.get(d.day)
            raw = row[col] if col is not None and col < len(row) else ""
            schedule[d][emp.key] = normalize_shift_code(raw, d, holidays)

    if not employees:
        raise ValueError("표에서 직원 행을 찾지 못했습니다.")

    expand_gy_blocks(employees, year, month, schedule)
    result = ScheduleResult(year=year, month=month, employees=employees, schedule=schedule, holidays=holidays)
    result.warnings = validate_schedule(employees, year, month, schedule, holidays, rules or ShiftRules())
    return result



def _rgb_from_openpyxl_color(color) -> Optional[tuple[int, int, int]]:
    if color is None:
        return None
    rgb = getattr(color, "rgb", None)
    if rgb and isinstance(rgb, str):
        rgb = rgb[-6:]
        try:
            return int(rgb[0:2], 16), int(rgb[2:4], 16), int(rgb[4:6], 16)
        except ValueError:
            return None
    indexed = getattr(color, "indexed", None)
    indexed_map = {
        22: (192, 192, 192),
        23: (128, 128, 128),
        48: (128, 128, 128),
        15: (192, 192, 192),
        16: (128, 128, 128),
    }
    if indexed in indexed_map:
        return indexed_map[indexed]
    return None


def is_gray_fill(cell) -> bool:
    fill = getattr(cell, "fill", None)
    if fill is None or not getattr(fill, "fill_type", None):
        return False
    rgb = _rgb_from_openpyxl_color(getattr(fill, "fgColor", None)) or _rgb_from_openpyxl_color(getattr(fill, "start_color", None))
    if rgb is None:
        # Theme colors are hard to resolve without workbook theme parsing. If a schedule cell
        # has any non-empty fill but no RGB, treat it as marked unavailable.
        fg = getattr(fill, "fgColor", None)
        if fg and getattr(fg, "type", None) == "theme":
            return True
        return False
    r, g, b = rgb
    if (r, g, b) in ((255, 255, 255), (0, 0, 0)):
        return False
    return max(r, g, b) - min(r, g, b) <= 18 and 70 <= (r + g + b) / 3 <= 235


def is_blue_fill(cell) -> bool:
    fill = getattr(cell, "fill", None)
    if fill is None or not getattr(fill, "fill_type", None):
        return False
    rgb = _rgb_from_openpyxl_color(getattr(fill, "fgColor", None)) or _rgb_from_openpyxl_color(getattr(fill, "start_color", None))
    return is_blue_rgb(rgb)


def import_schedule_from_excel(path: str | Path, year: int, month: int, rules: Optional[ShiftRules] = None) -> ScheduleResult:
    """Read an Excel roster directly, including cell fill colors.

    This is more reliable than clipboard paste for color-only G/지근 marks
    because openpyxl reads the workbook cell styles directly.
    """
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as exc:
        raise RuntimeError("openpyxl이 필요합니다. `pip install -r requirements.txt`를 실행하세요.") from exc

    wb = load_workbook(path, data_only=True)
    ws = wb.active
    lines: List[str] = []
    for row in ws.iter_rows():
        values: List[str] = []
        has_any = False
        for cell in row:
            value = "" if cell.value is None else str(cell.value).strip()
            if not value and is_blue_fill(cell):
                value = SHIFT_GY
            if value:
                has_any = True
            values.append(value)
        if has_any:
            lines.append("\t".join(values))
    if not lines:
        raise ValueError("엑셀 파일에서 근무표 데이터를 찾지 못했습니다.")
    return parse_schedule_from_tsv("\n".join(lines), year, month, rules)


def import_unavailable_from_gray_excel(path: str | Path, year: int, month: int) -> Dict[str, Set[date]]:
    """Read an Excel roster and return unavailable dates from gray-filled date cells.

    Expected sheet shape: 성명 | 사번 | 1 | 2 | ... | 말일.
    Keys are both employee key `name|employee_no` and employee_no/name fallbacks.
    """
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as exc:
        raise RuntimeError("openpyxl이 필요합니다. `pip install -r requirements.txt`를 실행하세요.") from exc

    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows())
    if not rows:
        return {}

    header_row_idx = 0
    name_col = 0
    id_col = 1
    for idx, row in enumerate(rows[:8]):
        values = [str(c.value).strip().replace(" ", "").lower() if c.value is not None else "" for c in row]
        if any(v in ("성명", "이름", "name") for v in values):
            header_row_idx = idx
            for col, v in enumerate(values):
                if v in ("성명", "이름", "name"):
                    name_col = col
                if v in ("사번", "직번", "id", "employeeid", "employee_id"):
                    id_col = col
            break

    valid_dates = {d.day: d for d in month_dates(year, month)}
    day_cols: Dict[int, int] = {}
    for col, cell in enumerate(rows[header_row_idx]):
        day = _header_day(cell.value)
        if day in valid_dates:
            day_cols[day] = col

    if not day_cols:
        for offset, d in enumerate(month_dates(year, month), start=2):
            day_cols[d.day] = offset

    result: Dict[str, Set[date]] = {}
    for row in rows[header_row_idx + 1:]:
        if name_col >= len(row) or row[name_col].value is None:
            continue
        name = str(row[name_col].value).strip()
        if not name:
            continue
        employee_no = str(row[id_col].value).strip() if id_col < len(row) and row[id_col].value is not None else ""
        dates: Set[date] = set()
        for day, col in day_cols.items():
            if col < len(row) and is_gray_fill(row[col]):
                dates.add(valid_dates[day])
        if dates:
            key = f"{name}|{employee_no}"
            result[key] = dates
            if employee_no:
                result[employee_no] = dates
            result[name] = dates
    return result

def import_employees_from_excel(path: str | Path) -> List[Employee]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as exc:
        raise RuntimeError("openpyxl이 필요합니다. `pip install -r requirements.txt`를 실행하세요.") from exc

    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(v).strip() if v is not None else "" for v in rows[0]]

    def find_col(names: Tuple[str, ...], default: Optional[int] = None) -> Optional[int]:
        for name in names:
            for i, h in enumerate(header):
                if name.lower() == h.lower():
                    return i
        return default

    name_col = find_col(("성명", "이름", "name"), 0)
    id_col = find_col(("사번", "employee_id", "id"), 1)
    new_col = find_col(("신규", "new", "is_new"), None)
    unavailable_col = find_col(("불가일", "근무불가", "unavailable", "unavailable_dates"), None)

    employees: List[Employee] = []
    for row in rows[1:]:
        if name_col is None or name_col >= len(row) or not row[name_col]:
            continue
        name = str(row[name_col]).strip()
        employee_id = str(row[id_col]).strip() if id_col is not None and id_col < len(row) and row[id_col] is not None else ""
        is_new = False
        if new_col is not None and new_col < len(row):
            is_new = str(row[new_col]).strip().lower() in ("y", "yes", "true", "1", "신규", "ㅇ", "o")
        unavailable = set()
        if unavailable_col is not None and unavailable_col < len(row):
            unavailable = parse_unavailable(row[unavailable_col])
        employees.append(Employee(name=name, employee_id=employee_id, is_new=is_new, unavailable_dates=unavailable))
    return employees


def parse_employees_from_tsv(text: str) -> List[Employee]:
    employees: List[Employee] = []
    for line in text.splitlines():
        if not line.strip():
            continue
        parts = [p.strip() for p in line.split("\t")]
        if parts[0] in ("성명", "이름", "name"):
            continue
        name = parts[0] if len(parts) >= 1 else ""
        employee_id = parts[1] if len(parts) >= 2 else ""
        is_new = len(parts) >= 3 and parts[2].lower() in ("y", "yes", "true", "1", "신규", "ㅇ", "o")
        unavailable = parse_unavailable(parts[3]) if len(parts) >= 4 else set()
        if name:
            employees.append(Employee(name=name, employee_id=employee_id, is_new=is_new, unavailable_dates=unavailable))
    return employees


def export_schedule_to_excel(result: ScheduleResult, path: str | Path) -> None:
    try:
        from openpyxl import Workbook  # type: ignore
        from openpyxl.styles import Alignment, Font, PatternFill, Border, Side  # type: ignore
        from openpyxl.utils import get_column_letter  # type: ignore
    except Exception as exc:
        raise RuntimeError("openpyxl이 필요합니다. `pip install -r requirements.txt`를 실행하세요.") from exc

    wb = Workbook()
    ws = wb.active
    ws.title = "근무표"
    stats_ws = wb.create_sheet("통계")
    warn_ws = wb.create_sheet("검증")

    dates = month_dates(result.year, result.month)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.cell(row=1, column=1, value="성명")
    ws.cell(row=1, column=2, value="사번")
    for idx, d in enumerate(dates, start=3):
        c = ws.cell(row=1, column=idx, value=weekday_ko(d))
        c.alignment = Alignment(horizontal="center")
        c.font = Font(bold=True)
        c = ws.cell(row=2, column=idx, value=d.day)
        c.alignment = Alignment(horizontal="center")
        if is_holiday_or_weekend(d, result.holidays):
            c.fill = PatternFill("solid", fgColor="F4CCCC")

    for row_idx, emp in enumerate(result.employees, start=3):
        ws.cell(row=row_idx, column=1, value=emp.name)
        ws.cell(row=row_idx, column=2, value=emp.employee_id)
        for col_idx, d in enumerate(dates, start=3):
            shift = result.schedule.get(d, {}).get(emp.key, OFF)
            cell = ws.cell(row=row_idx, column=col_idx, value=shift)
            cell.alignment = Alignment(horizontal="center")
            cell.fill = PatternFill("solid", fgColor=SHIFT_FILLS.get(shift, "FFFFFF"))
            cell.border = border

    for col in range(1, len(dates) + 3):
        ws.column_dimensions[get_column_letter(col)].width = 8 if col >= 3 else 14
    ws.freeze_panes = "C3"

    stats = compute_stats(result.employees, dates, result.schedule, result.holidays)
    avg = averages(stats)
    for col, header in enumerate(STAT_HEADERS, start=1):
        stats_ws.cell(row=1, column=col, value=header).font = Font(bold=True)
    stats_ws.cell(row=1, column=len(STAT_HEADERS) + 1, value="총근무 평균편차").font = Font(bold=True)
    for row_idx, stat in enumerate(stats.values(), start=2):
        for col_idx, value in enumerate(stat.as_row(), start=1):
            stats_ws.cell(row=row_idx, column=col_idx, value=value)
        stats_ws.cell(row=row_idx, column=len(STAT_HEADERS) + 1, value=round(stat.total_work - avg.get("total_work", 0), 2))
    for col in range(1, len(STAT_HEADERS) + 2):
        stats_ws.column_dimensions[get_column_letter(col)].width = 14

    warn_ws.cell(row=1, column=1, value="경고")
    warn_ws.cell(row=1, column=1).font = Font(bold=True)
    if result.warnings:
        for row_idx, warning in enumerate(result.warnings, start=2):
            warn_ws.cell(row=row_idx, column=1, value=warning)
    else:
        warn_ws.cell(row=2, column=1, value="검증 경고 없음")
    warn_ws.column_dimensions["A"].width = 80

    wb.save(path)
